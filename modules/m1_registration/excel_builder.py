import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

from config.settings import OUTPUT_DIR
from core.logger import get_logger
from modules.m1_registration.mapper import map_to_platform, validate_mapped

logger = get_logger("excel_builder")

# 플랫폼별 헤더 색상
PLATFORM_COLORS = {
    "smartstore": "1D6340",   # 스마트스토어 초록
    "aladin":     "E8534A",   # 알라딘 빨강
    "yes24":      "E8860A",   # 예스24 주황
}


def _apply_header_style(ws, platform: str):
    """헤더 행 스타일 적용"""
    color = PLATFORM_COLORS.get(platform, "2E75B6")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22


def _apply_data_style(ws):
    """데이터 행 스타일 적용"""
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = "FFFFFF" if row_idx % 2 == 0 else "F7F9FC"
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[row_idx].height = 18


def _auto_column_width(ws):
    """컬럼 너비 자동 조정"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                # 한글은 2배 너비
                kor_count = sum(1 for c in str(cell.value or "") if ord(c) > 127)
                val_len += kor_count
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def build_upload_excel(
    df: pd.DataFrame,
    platform: str,
    output_dir: Path = None
) -> Path:
    """
    플랫폼별 업로드용 엑셀 파일 생성
    반환: 생성된 파일 경로
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    # 매핑
    mapped_df = map_to_platform(df, platform)

    # 유효성 검사
    errors = validate_mapped(mapped_df, platform)
    if errors:
        for err in errors:
            logger.warning(f"[{platform}] 유효성 경고: {err}")

    # 파일명: 플랫폼_날짜시간.xlsx
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{platform}_upload_{timestamp}.xlsx"
    output_path = output_dir / filename

    # 엑셀 작성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{platform}_업로드"

    # 헤더 행 작성
    for col_idx, col_name in enumerate(mapped_df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 데이터 행 작성
    for row_idx, row in enumerate(mapped_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 스타일 적용
    _apply_header_style(ws, platform)
    _apply_data_style(ws)
    _auto_column_width(ws)

    # 헤더 고정
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"[{platform}] 업로드 파일 생성: {output_path.name} ({len(mapped_df)}건)")

    return output_path


def build_all_platforms(df: pd.DataFrame, output_dir: Path = None) -> dict:
    """
    등록 대상 플랫폼 자동 감지 후 전체 생성
    반환: {플랫폼명: 파일경로}
    """
    results = {}
    platform_col_map = {
        "smartstore": "등록대상_스마트",
        "aladin":     "등록대상_알라딘",
        "yes24":      "등록대상_예스24",
    }

    for platform, col in platform_col_map.items():
        if col not in df.columns:
            logger.warning(f"[{platform}] 컬럼 없음, 건너뜀: {col}")
            continue
        target_df = df[df[col].str.upper() == "Y"].copy()
        if target_df.empty:
            logger.info(f"[{platform}] 등록 대상 없음, 건너뜀")
            continue
        try:
            path = build_upload_excel(target_df, platform, output_dir)
            results[platform] = path
        except Exception as e:
            logger.error(f"[{platform}] 파일 생성 실패: {e}")

    return results


if __name__ == "__main__":
    # 샘플 데이터로 테스트
    sample_df = pd.DataFrame([
        {
            "상품ID": "P001", "ISBN": "9788936434267",
            "도서명": "채식주의자", "도서명_조합": "",
            "저자": "한강", "출판사": "창비",
            "판매가": "6000", "상태등급": "상", "재고수량": "1",
            "등록대상_스마트": "Y", "등록대상_알라딘": "Y", "등록대상_예스24": "N",
        },
        {
            "상품ID": "P002", "ISBN": "9788954651135",
            "도서명": "82년생 김지영", "도서명_조합": "82년생 김지영",
            "저자": "조남주", "출판사": "민음사",
            "판매가": "8000", "상태등급": "최상", "재고수량": "1",
            "등록대상_스마트": "Y", "등록대상_알라딘": "Y", "등록대상_예스24": "Y",
        },
    ])

    print("=== excel_builder.py 테스트 ===")

    # 단일 플랫폼 테스트
    path = build_upload_excel(sample_df, "aladin")
    print(f"알라딘 파일: {path.name}")

    # 전체 플랫폼 자동 생성 테스트
    results = build_all_platforms(sample_df)
    print(f"\n전체 생성 결과:")
    for platform, path in results.items():
        print(f"  [{platform}] {path.name}")

    # output 폴더 확인
    print(f"\noutput/ 폴더:")
    for f in sorted(OUTPUT_DIR.iterdir()):
        size = f.stat().st_size
        print(f"  {f.name} ({size:,} bytes)")

    print("\n=== 테스트 완료 ===")
