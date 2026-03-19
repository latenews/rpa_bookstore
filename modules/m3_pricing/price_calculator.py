import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import PRICING, OUTPUT_DIR
from core.logger import get_logger
from modules.m3_pricing.crawler import get_aladin_min_price_bulk

logger = get_logger("price_calculator")


def calculate_adjusted_price(
    my_price: int,
    competitor_min: int,
    min_allowed: int,
    undercut: int = None,
    min_gap: int = None,
) -> tuple:
    """
    조정 가격 계산 및 예외처리
    반환: (조정가 또는 None, 사유 문자열)
    """
    undercut  = undercut  or PRICING["undercut_amount"]
    min_gap   = min_gap   or PRICING["min_gap"]

    # 예외 1: 경쟁가 없음
    if competitor_min is None:
        return None, "경쟁가 조회 실패"

    # 예외 2: 내 가격이 이미 최저가 이하
    if my_price <= competitor_min:
        return None, f"이미 최저가 이하 (내:{my_price:,} / 경쟁:{competitor_min:,})"

    # 예외 3: 가격 차이가 min_gap 이하 → 미조정
    gap = my_price - competitor_min
    if gap <= min_gap:
        return None, f"가격차 {gap:,}원 (기준 {min_gap:,}원 이하, 미조정)"

    # 조정 목표가
    target = competitor_min - undercut

    # 예외 4: 목표가가 최저허용가 이하
    if target < min_allowed:
        return None, f"목표가 {target:,}원이 최저허용가 {min_allowed:,}원 미만"

    return target, f"경쟁최저 {competitor_min:,}원 - {undercut:,}원 = {target:,}원"


def run_price_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """
    1단계: 가격 비교 및 조정 대상 분석
    반환: 분석 결과 DataFrame
    """
    logger.info(f"가격 분석 시작: {len(df)}건")

    # ISBN 목록으로 일괄 크롤링
    isbn_list = df["ISBN"].tolist()
    crawl_results = get_aladin_min_price_bulk(isbn_list)

    rows = []
    for _, row in df.iterrows():
        isbn        = row.get("ISBN", "")
        my_price    = int(row.get("판매가", 0) or 0)
        min_allowed = int(row.get("최저허용가", PRICING["min_allowed_price"]) or PRICING["min_allowed_price"])
        comp_min    = crawl_results.get(isbn)

        adjusted, reason = calculate_adjusted_price(my_price, comp_min, min_allowed)

        rows.append({
            "상품ID":       row.get("상품ID", ""),
            "ISBN":         isbn,
            "도서명":       row.get("도서명", ""),
            "현재판매가":   my_price,
            "경쟁최저가":   comp_min if comp_min else "-",
            "최저허용가":   min_allowed,
            "조정가":       adjusted if adjusted else "-",
            "조정여부":     "Y" if adjusted else "N",
            "사유":         reason,
        })

    result_df = pd.DataFrame(rows)
    adj_count = (result_df["조정여부"] == "Y").sum()
    logger.info(f"분석 완료: 전체 {len(result_df)}건 / 조정대상 {adj_count}건")
    return result_df


def export_price_report(result_df: pd.DataFrame) -> Path:
    """
    1차 산출물: 가격 비교 보고서 엑셀 출력
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = OUTPUT_DIR / f"price_report_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가격분석"

    # 컬럼 색상 정의
    col_colors = {
        "상품ID":     "D9D9D9",
        "ISBN":       "D9D9D9",
        "도서명":     "D9D9D9",
        "현재판매가": "BDD7EE",
        "경쟁최저가": "FCE4D6",
        "최저허용가": "E2EFDA",
        "조정가":     "FFF2CC",
        "조정여부":   "FFF2CC",
        "사유":       "F2F2F2",
    }

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    for col_idx, col_name in enumerate(result_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        bg = col_colors.get(col_name, "D9D9D9")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 데이터
    for row_idx, row in enumerate(result_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            # 조정여부 Y → 노란 배경 강조
            col_name = result_df.columns[col_idx - 1]
            if col_name == "조정여부" and value == "Y":
                cell.fill = PatternFill("solid", fgColor="FFEB3B")
                cell.font = Font(bold=True, size=10, color="C00000")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            kor = sum(1 for c in val if ord(c) > 127)
            max_len = max(max_len, len(val) + kor)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    adj_df = result_df[result_df["조정여부"] == "Y"]
    total   = len(result_df)
    adj     = len(adj_df)
    skip    = total - adj

    summary = [
        ["항목", "건수"],
        ["전체 분석", total],
        ["조정 대상", adj],
        ["미조정 (예외처리)", skip],
        ["", ""],
        ["생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for r, row in enumerate(summary, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    logger.info(f"가격 보고서 저장: {path.name}")
    return path


if __name__ == "__main__":
    print("=== price_calculator.py 테스트 ===\n")

    # 단위 테스트
    print("[ 가격 계산 로직 단위 테스트 ]")
    cases = [
        (8000, 7000, 3000, "정상 조정"),
        (8000, 8500, 3000, "이미 최저가"),
        (7050, 7000, 3000, "가격차 미미"),
        (4000, 3500, 3000, "목표가 최저허용가 미만"),
        (8000, None, 3000, "경쟁가 없음"),
    ]
    for my, comp, minp, desc in cases:
        adj, reason = calculate_adjusted_price(my, comp, minp)
        result = f"{adj:,}원" if adj else "미조정"
        print(f"  [{desc}] 내:{my:,} / 경쟁:{comp} → {result} ({reason})")

    # 실제 크롤링 통합 테스트
    print("\n[ 실제 크롤링 통합 테스트 ]")
    sample_df = pd.DataFrame([
        {"상품ID": "P001", "ISBN": "9788936434267",
         "도서명": "채식주의자",    "판매가": "8000", "최저허용가": "2000"},
        {"상품ID": "P002", "ISBN": "9788954651135",
         "도서명": "82년생 김지영", "판매가": "12000", "최저허용가": "3000"},
    ])

    result_df = run_price_analysis(sample_df)
    print(result_df.to_string(index=False))

    path = export_price_report(result_df)
    print(f"\n✅ 보고서 저장: {path.name}")
    print("=== 테스트 완료 ===")
