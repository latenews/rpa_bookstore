import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

from config.settings import OUTPUT_DIR
from core.logger import get_logger

logger = get_logger("invoice_handler")

# 택배사별 송장 양식 컬럼
COURIER_COLUMNS = {
    "CJ대한통운": [
        "받는분성명", "받는분전화번호", "받는분주소",
        "품목", "수량", "무게(kg)", "박스수량", "배송메모",
    ],
    "한진택배": [
        "수하인명", "수하인전화", "수하인주소",
        "상품명", "수량", "무게", "메모",
    ],
    "우체국택배": [
        "받는사람", "전화번호", "주소",
        "내용물", "수량", "메모",
    ],
}

# 표준 → 택배사 컬럼 매핑
COURIER_MAPPING = {
    "CJ대한통운": {
        "receiver_name":  "받는분성명",
        "receiver_phone": "받는분전화번호",
        "address":        "받는분주소",
        "product_name":   "품목",
        "quantity":       "수량",
        "memo":           "배송메모",
    },
    "한진택배": {
        "receiver_name":  "수하인명",
        "receiver_phone": "수하인전화",
        "address":        "수하인주소",
        "product_name":   "상품명",
        "quantity":       "수량",
        "memo":           "메모",
    },
    "우체국택배": {
        "receiver_name":  "받는사람",
        "receiver_phone": "전화번호",
        "address":        "주소",
        "product_name":   "내용물",
        "quantity":       "수량",
        "memo":           "메모",
    },
}


def build_invoice_excel(merged_df: pd.DataFrame, courier: str = "CJ대한통운") -> Path:
    """
    통합 주문 DataFrame → 택배사 송장 출력용 엑셀 생성
    """
    if courier not in COURIER_MAPPING:
        raise ValueError(f"지원하지 않는 택배사: {courier}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = OUTPUT_DIR / f"invoice_{courier}_{timestamp}.xlsx"

    mapping = COURIER_MAPPING[courier]
    courier_cols = list(mapping.values())

    # 데이터 변환
    rows = []
    for _, row in merged_df.iterrows():
        new_row = {}
        for std_col, courier_col in mapping.items():
            new_row[courier_col] = row.get(std_col, "")
        rows.append(new_row)

    df_out = pd.DataFrame(rows, columns=courier_cols)

    # 스타일 적용해서 저장
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{courier}_송장"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    for col_idx, col_name in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    # 데이터
    for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF3FB")
        ws.row_dimensions[row_idx].height = 20

    # 컬럼 너비
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            kor = sum(1 for c in val if ord(c) > 127)
            max_len = max(max_len, len(val) + kor)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(path)
    logger.info(f"[{courier}] 송장 파일 생성: {path.name} ({len(df_out)}건)")
    return path


def update_invoice_numbers(merged_df: pd.DataFrame, invoice_data: list) -> pd.DataFrame:
    """
    송장번호 일괄 업데이트
    invoice_data: [{"order_id": ..., "invoice_no": ..., "courier": ...}, ...]
    """
    df = merged_df.copy()
    updated = 0
    for item in invoice_data:
        mask = df["order_id"] == item["order_id"]
        if mask.sum() > 0:
            df.loc[mask, "invoice_no"] = item.get("invoice_no", "")
            df.loc[mask, "courier"]    = item.get("courier", "")
            df.loc[mask, "status"]     = "송장등록"
            updated += 1
    logger.info(f"송장번호 업데이트: {updated}건")
    return df


def export_merged_orders(merged_df: pd.DataFrame) -> Path:
    """
    통합 주문 현황 엑셀 저장 (송장번호 포함)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = OUTPUT_DIR / f"orders_merged_{timestamp}.xlsx"

    # 한글 컬럼명으로 변환
    col_kr = {
        "order_id":       "주문번호",
        "platform":       "플랫폼",
        "order_date":     "주문일시",
        "receiver_name":  "수취인",
        "receiver_phone": "연락처",
        "address":        "배송주소",
        "product_name":   "상품명",
        "quantity":       "수량",
        "amount":         "금액",
        "memo":           "배송메모",
        "invoice_no":     "송장번호",
        "courier":        "택배사",
        "status":         "처리상태",
    }
    df_out = merged_df.rename(columns=col_kr)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "통합주문현황"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_colors = {
        "미처리":   "FCE4D6",
        "송장등록": "E2EFDA",
        "발송완료": "BDD7EE",
    }

    for col_idx, col_name in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
        status = str(row[-1]) if len(row) > 0 else "미처리"
        row_bg = status_colors.get(status, "FFFFFF")
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[row_idx].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            kor = sum(1 for c in val if ord(c) > 127)
            max_len = max(max_len, len(val) + kor)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(path)
    logger.info(f"통합 주문 현황 저장: {path.name}")
    return path


if __name__ == "__main__":
    from modules.m4_orders.order_merger import merge_orders, create_sample_order_files
    from config.settings import OUTPUT_DIR

    print("=== invoice_handler.py 테스트 ===")

    sample_dir = OUTPUT_DIR / "sample_orders"
    file_list = create_sample_order_files(sample_dir)
    merged = merge_orders(file_list)
    print(f"주문 통합: {len(merged)}건")

    # 송장 파일 생성
    for courier in ["CJ대한통운", "한진택배", "우체국택배"]:
        path = build_invoice_excel(merged, courier)
        print(f"✅ [{courier}] {path.name}")

    # 송장번호 업데이트 테스트
    invoice_data = [
        {"order_id": "SS-20260319-001", "invoice_no": "123456789012", "courier": "CJ대한통운"},
        {"order_id": "AL-20260319-001", "invoice_no": "987654321098", "courier": "한진택배"},
    ]
    updated = update_invoice_numbers(merged, invoice_data)
    print(f"\n송장번호 업데이트 후:")
    print(updated[["order_id", "invoice_no", "courier", "status"]].to_string(index=False))

    # 통합 주문 현황 저장
    path = export_merged_orders(updated)
    print(f"\n✅ 통합 주문 현황: {path.name}")
    print("=== 테스트 완료 ===")
