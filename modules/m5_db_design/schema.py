import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from config.settings import TEMPLATE_DIR
from core.logger import get_logger

logger = get_logger("schema")

# 컬럼 정의: (컬럼명, 너비, 설명, 필수여부)
SCHEMA = [
    ("상품ID",          12, "내부 고유키 (자동생성)",       True),
    ("ISBN",            16, "13자리 ISBN",                  True),
    ("도서명",          30, "원제목",                       True),
    ("도서명_조합",     35, "권수 포함 조합명 (자동생성)",  False),
    ("저자",            15, "저자명",                       False),
    ("출판사",          15, "출판사명",                     False),
    ("정가",            10, "정가 (숫자)",                  False),
    ("판매가",          10, "희망 판매가 (숫자)",           True),
    ("최저허용가",      12, "가격조정 최저 허용가",         False),
    ("상태등급",        10, "최상/상/중/하",                True),
    ("재고수량",        10, "재고 수량",                    True),
    ("등록대상_스마트", 14, "스마트스토어 등록 Y/N",        False),
    ("등록대상_알라딘", 14, "알라딘 등록 Y/N",             False),
    ("등록대상_예스24", 14, "예스24 등록 Y/N",             False),
    ("알라딘_상품코드", 16, "알라딘 등록 후 코드",          False),
    ("개똥이네_상품코드",16,"개똥이네 등록 후 코드",        False),
    ("북코아_상품코드", 16, "북코아 등록 후 코드",          False),
    ("판매완료여부",    12, "판매 완료 Y/N",                False),
    ("판매채널",        12, "판매된 채널명",                False),
    ("처리상태",        10, "미처리/처리중/완료",           False),
]

# 색상 정의
COLOR = {
    "header_required": "C00000",   # 진빨강 — 필수컬럼
    "header_optional": "2E75B6",   # 파랑   — 선택컬럼
    "header_font":     "FFFFFF",   # 흰색 글자
    "desc_bg":         "F2F2F2",   # 설명행 배경
    "sample_bg":       "EBF3FB",   # 샘플행 배경
    "alt_row":         "FAFAFA",   # 짝수행 배경
}

SAMPLE_DATA = [
    ["P001","9788936434267","채식주의자","채식주의자","한강","창비",
     11000,6000,3000,"상",1,"Y","Y","N","","","","N","","미처리"],
    ["P002","9788954651135","82년생 김지영","82년생 김지영","조남주","민음사",
     14000,8000,4000,"최상",1,"Y","Y","Y","ALI-12345","","","N","","미처리"],
]


def create_master_template(output_path: Path = None) -> Path:
    """마스터 DB 엑셀 템플릿 생성"""
    if output_path is None:
        TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = TEMPLATE_DIR / "master_db_template.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "마스터DB"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 1행: 설명 행 ──
    for col_idx, (name, width, desc, required) in enumerate(SCHEMA, start=1):
        cell = ws.cell(row=1, column=col_idx, value=desc)
        cell.fill = PatternFill("solid", fgColor=COLOR["desc_bg"])
        cell.font = Font(size=9, color="666666", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 2행: 헤더 행 ──
    for col_idx, (name, width, desc, required) in enumerate(SCHEMA, start=1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        bg = COLOR["header_required"] if required else COLOR["header_optional"]
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(size=11, bold=True, color=COLOR["header_font"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── 3행~: 샘플 데이터 ──
    for row_idx, row_data in enumerate(SAMPLE_DATA, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=COLOR["sample_bg"])
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # 행 높이 설정
    ws.row_dimensions[1].height = 30   # 설명행
    ws.row_dimensions[2].height = 25   # 헤더행
    for i in range(3, 3 + len(SAMPLE_DATA)):
        ws.row_dimensions[i].height = 20

    # 헤더 고정 (2행 아래부터 스크롤)
    ws.freeze_panes = "A3"

    # 범례 시트 추가
    ws2 = wb.create_sheet("사용법")
    guide = [
        ["항목",        "설명"],
        ["상품ID",      "비워두세요. 시스템이 자동 생성합니다."],
        ["상태등급",    "최상 / 상 / 중 / 하 중 하나 입력"],
        ["등록대상_*",  "등록할 플랫폼은 Y, 아니면 N 입력"],
        ["판매완료여부","판매 완료 시 Y로 변경 → 모듈2가 자동 처리"],
        ["처리상태",    "미처리 / 처리중 / 완료 (시스템이 자동 변경)"],
        ["빨간 헤더",   "필수 입력 항목"],
        ["파란 헤더",   "선택 입력 항목"],
    ]
    for r, row in enumerate(guide, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    wb.save(output_path)
    logger.info(f"템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == "__main__":
    path = create_master_template()
    print(f"✅ 템플릿 생성: {path}")
    print(f"   컬럼 수: {len(SCHEMA)}개")
    print(f"   필수 컬럼: {sum(1 for _,_,_,r in SCHEMA if r)}개")
    print(f"   선택 컬럼: {sum(1 for _,_,_,r in SCHEMA if not r)}개")
